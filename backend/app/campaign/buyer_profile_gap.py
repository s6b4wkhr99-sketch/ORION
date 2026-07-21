"""Buyer profile matrix + bias-adjusted GAP analysis (read-only, no dashboard impact)."""

from __future__ import annotations

import csv
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl

from app.intelligence.buyer_gap_mapping import parse_purchase_token, purchase_series

US_STATES = frozenset(
    {
        "AL", "AK", "AZ", "AR", "CA", "CO", "CT", "DE", "FL", "GA", "HI", "ID", "IL", "IN", "IA",
        "KS", "KY", "LA", "ME", "MD", "MA", "MI", "MN", "MS", "MO", "MT", "NE", "NV", "NH", "NJ",
        "NM", "NY", "NC", "ND", "OH", "OK", "OR", "PA", "RI", "SC", "SD", "TN", "TX", "UT", "VT",
        "VA", "WA", "WV", "WI", "WY", "DC",
    }
)

PRIORITY_STATES = frozenset({"TX", "FL", "NY"})

MASTER_TO_TOKEN = {
    "Master S4": "V4",
    "Master V5": "V5",
    "Master V6": "V6",
    "Master V7": "V7",
    "Master V9": "V9",
    "Pause M4": "M4",
    "Pause M6": "M6",
    "Pause M6s": "M6S",
    "Pause M10": "M10",
}

TOKEN_TO_MASTER = {v: k for k, v in MASTER_TO_TOKEN.items()}

_STATE_IN_TEXT = re.compile(
    r"\b(" + "|".join(sorted(US_STATES, key=len, reverse=True)) + r")\b"
)


@dataclass
class BuyerRow:
    email: str
    product_raw: str
    sku_token: str
    source: str
    row_id: str
    state: str
    state_tier: str
    era: str
    series: str | None


def norm_email(value: str | None) -> str | None:
    email = (value or "").strip().lower()
    if not email or "@" not in email:
        return None
    return email


def parse_us_state(raw: str | None, *, source: str) -> str:
    text = (raw or "").strip().upper()
    if not text:
        return "OTHER"
    if source == "legacy":
        match = _STATE_IN_TEXT.search(text)
        if match:
            return match.group(1)
        if text.startswith("CA") or " CALIF" in text:
            return "CA"
        return "OTHER"
    code = text[:2] if len(text) >= 2 else text
    return code if code in US_STATES else "OTHER"


def state_tier(state: str) -> str:
    if state == "CA":
        return "CA"
    if state in PRIORITY_STATES:
        return "PRIORITY"
    if state in US_STATES:
        return "REST_US"
    return "OTHER"


def parse_era(paid_at: str | None, *, source: str) -> str:
    if source == "legacy":
        return "pre2025"
    paid = (paid_at or "").strip()
    if paid.startswith("2025") or paid.startswith("2026"):
        return "post2025"
    return "pre2025"


def pct(count: int, total: int) -> float:
    return round(100.0 * count / total, 2) if total else 0.0


def counter_to_pct(counter: Counter[str], total: int) -> dict[str, float]:
    if not total:
        return {}
    return {k: round(100.0 * counter[k] / total, 2) for k in sorted(counter, key=lambda x: -counter[x])}


def load_buyer_rows(legacy: Path, shopify: Path) -> list[BuyerRow]:
    rows: list[BuyerRow] = []
    wb = openpyxl.load_workbook(legacy, read_only=True, data_only=True)
    ws = wb["SALES (202112-202404)"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(headers)}
    for n, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        status = str(row[idx.get("Status", 0)] or "")
        if status.upper() not in ("PAID", "PROCESSING"):
            continue
        email = norm_email(row[idx.get("E MAIL", 0)])
        prod = str(row[idx.get("Material Name", 0)] or "")
        token = parse_purchase_token(prod)
        loc = str(row[idx.get("LOCATION", 0)] or "")
        state = parse_us_state(loc, source="legacy")
        if email and token:
            rows.append(
                BuyerRow(
                    email=email,
                    product_raw=prod,
                    sku_token=token,
                    source="legacy",
                    row_id=f"legacy:{n}",
                    state=state,
                    state_tier=state_tier(state),
                    era="pre2025",
                    series=purchase_series(token),
                )
            )
    wb.close()

    with open(shopify, newline="", encoding="utf-8-sig") as handle:
        for n, row in enumerate(csv.DictReader(handle), start=2):
            if (row.get("Financial Status") or "").lower() != "paid":
                continue
            email = norm_email(row.get("Email"))
            prod = row.get("Lineitem name") or row.get("Lineitem sku") or ""
            token = parse_purchase_token(prod)
            ship = row.get("Shipping Province") or row.get("Billing Province") or ""
            state = parse_us_state(ship, source="shopify")
            era = parse_era(row.get("Paid at"), source="shopify")
            if email and token:
                rows.append(
                    BuyerRow(
                        email=email,
                        product_raw=prod,
                        sku_token=token,
                        source="shopify",
                        row_id=f"shopify:{n}",
                        state=state,
                        state_tier=state_tier(state),
                        era=era,
                        series=purchase_series(token),
                    )
                )
    return rows


def build_profile_matrix(rows: list[BuyerRow]) -> dict[str, Any]:
    """SKU distributions by geography tier, era, channel, and cross-tabs."""

    def profile_for(subset: list[BuyerRow]) -> dict[str, Any]:
        n = len(subset)
        sku = Counter(r.sku_token for r in subset)
        series = Counter(r.series for r in subset if r.series)
        return {
            "rows": n,
            "unique_emails": len({r.email for r in subset}),
            "sku_counts": dict(sku),
            "sku_pct": counter_to_pct(sku, n),
            "series_pct": counter_to_pct(series, n),
            "top_sku": sku.most_common(3),
        }

    matrix: dict[str, Any] = {"total_rows": len(rows), "profiles": {}}

    # Single-dimension buckets
    for dim, getter in [
        ("state_tier", lambda r: r.state_tier),
        ("era", lambda r: r.era),
        ("channel", lambda r: r.source),
        ("series_purchased", lambda r: r.series or "unknown"),
    ]:
        buckets: dict[str, list[BuyerRow]] = defaultdict(list)
        for row in rows:
            buckets[getter(row)].append(row)
        matrix["profiles"][dim] = {k: profile_for(v) for k, v in sorted(buckets.items())}

    # State-level (US only, min rows threshold noted in output)
    by_state: dict[str, list[BuyerRow]] = defaultdict(list)
    for row in rows:
        if row.state in US_STATES:
            by_state[row.state].append(row)
    matrix["profiles"]["by_state"] = {
        st: profile_for(subset) for st, subset in sorted(by_state.items(), key=lambda x: -len(x[1]))
    }

    # Cross: state_tier × era × channel
    cross: dict[str, list[BuyerRow]] = defaultdict(list)
    for row in rows:
        key = f"{row.state_tier}|{row.era}|{row.source}"
        cross[key].append(row)
    matrix["profiles"]["state_tier_x_era_x_channel"] = {
        k: profile_for(v) for k, v in sorted(cross.items())
    }

    return matrix


def compute_reweight_table(
    buyer_state_counts: Counter[str],
    prospect_state_counts: Counter[str],
) -> dict[str, Any]:
    """Prospect-share / buyer-share weights to debias geographic over/under-index."""
    buyer_total = sum(buyer_state_counts.values()) or 1
    prospect_total = sum(prospect_state_counts.values()) or 1

    buyer_us = sum(buyer_state_counts[s] for s in US_STATES)
    prospect_us = sum(prospect_state_counts.get(s, 0) for s in US_STATES)

    rows: list[dict[str, Any]] = []
    for state in sorted(US_STATES):
        b = buyer_state_counts.get(state, 0)
        p = prospect_state_counts.get(state, 0)
        if b == 0 and p == 0:
            continue
        buyer_share = b / buyer_us if buyer_us else 0
        prospect_share = p / prospect_us if prospect_us else 0
        weight = round(prospect_share / buyer_share, 4) if buyer_share > 0 else None
        rows.append(
            {
                "state": state,
                "buyer_rows": b,
                "buyer_share_pct": round(100 * buyer_share, 2),
                "prospect_rows": p,
                "prospect_share_pct": round(100 * prospect_share, 2),
                "reweight": weight,
                "over_index_in_buyers": round(buyer_share / prospect_share, 2)
                if prospect_share > 0 and buyer_share > 0
                else None,
            }
        )

    tier_buyer = Counter()
    tier_prospect = Counter()
    for state in US_STATES:
        tier = state_tier(state)
        tier_buyer[tier] += buyer_state_counts.get(state, 0)
        tier_prospect[tier] += prospect_state_counts.get(state, 0)

    tier_table: dict[str, Any] = {}
    for tier in ("CA", "PRIORITY", "REST_US"):
        b = tier_buyer[tier]
        p = tier_prospect[tier]
        b_share = b / buyer_us if buyer_us else 0
        p_share = p / prospect_us if prospect_us else 0
        tier_table[tier] = {
            "buyer_rows": b,
            "buyer_share_pct": round(100 * b_share, 2),
            "prospect_rows": p,
            "prospect_share_pct": round(100 * p_share, 2),
            "reweight": round(p_share / b_share, 4) if b_share > 0 else None,
            "over_index_in_buyers": round(b_share / p_share, 2) if p_share > 0 and b_share > 0 else None,
        }

    ca = tier_table.get("CA", {})
    ca_bias = ca.get("over_index_in_buyers")
    if ca_bias is None and ca.get("prospect_share_pct"):
        ca_bias = round(ca["buyer_share_pct"] / max(ca["prospect_share_pct"], 0.01), 2)

    return {
        "buyer_us_rows": buyer_us,
        "prospect_us_rows": prospect_us,
        "by_state": rows,
        "by_state_tier": tier_table,
        "ca_bias_index": ca_bias,
    }


def reweighted_buyer_sku_distribution(
    rows: list[BuyerRow],
    reweight_by_state: dict[str, float],
) -> dict[str, float]:
    """Each buyer row weighted by inverse geographic bias (prospect/buyer state share)."""
    weighted: Counter[str] = Counter()
    total_w = 0.0
    for row in rows:
        if row.state not in US_STATES:
            w = reweight_by_state.get("OTHER", 0.5)
        else:
            w = reweight_by_state.get(row.state, 1.0)
        weighted[row.sku_token] += w
        total_w += w
    if not total_w:
        return {}
    return {k: round(100.0 * weighted[k] / total_w, 2) for k in sorted(weighted, key=lambda x: -weighted[x])}


def prospect_token_distribution(prospect_rec_counts: dict[str, int]) -> dict[str, float]:
    total = sum(prospect_rec_counts.values()) or 1
    token_dist: Counter[str] = Counter()
    for master, n in prospect_rec_counts.items():
        token = MASTER_TO_TOKEN.get(master)
        if token:
            token_dist[token] += n
    return {k: round(100.0 * token_dist[k] / total, 2) for k in sorted(token_dist, key=lambda x: -token_dist[x])}


def distribution_gap(buyer_pct: dict[str, float], prospect_pct: dict[str, float]) -> dict[str, dict[str, float]]:
    tokens = sorted(set(buyer_pct) | set(prospect_pct))
    out: dict[str, dict[str, float]] = {}
    for token in tokens:
        b = buyer_pct.get(token, 0.0)
        p = prospect_pct.get(token, 0.0)
        out[token] = {
            "buyer_pct": b,
            "prospect_pct": p,
            "gap_points": round(b - p, 2),
        }
    return out


def expected_sku_for_prospect_state(
    state: str,
    profile_matrix: dict[str, Any],
    *,
    min_state_rows: int = 30,
) -> dict[str, float]:
    """Pick buyer SKU profile for a prospect state with fallback hierarchy."""
    by_state = profile_matrix["profiles"]["by_state"]
    if state in by_state and by_state[state]["rows"] >= min_state_rows:
        return by_state[state]["sku_pct"]

    tier = state_tier(state)
    tier_profiles = profile_matrix["profiles"]["state_tier"]
    if tier in tier_profiles and tier_profiles[tier]["rows"] > 0:
        return tier_profiles[tier]["sku_pct"]

    return profile_matrix["profiles"]["state_tier"].get("REST_US", {}).get("sku_pct", {})


def build_prospect_state_gap(
    prospect_by_state_rec: dict[str, Counter[str]],
    profile_matrix: dict[str, Any],
) -> dict[str, Any]:
    """Per-state: prospect recommended mix vs buyer profile for that state."""
    results = []
    for state, rec_counter in sorted(prospect_by_state_rec.items(), key=lambda x: -sum(x[1].values())):
        if state not in US_STATES:
            continue
        total = sum(rec_counter.values()) or 1
        prospect_pct = prospect_token_distribution(dict(rec_counter))
        buyer_pct = expected_sku_for_prospect_state(state, profile_matrix)
        gap = distribution_gap(buyer_pct, prospect_pct)
        # Largest absolute gaps
        top_gaps = sorted(gap.items(), key=lambda x: abs(x[1]["gap_points"]), reverse=True)[:3]
        results.append(
            {
                "state": state,
                "prospect_rows": total,
                "buyer_profile_rows": profile_matrix["profiles"]["by_state"].get(state, {}).get("rows", 0),
                "buyer_profile_source": (
                    "state"
                    if profile_matrix["profiles"]["by_state"].get(state, {}).get("rows", 0) >= 30
                    else state_tier(state)
                ),
                "top_gaps": [
                    {"sku": sku, **vals} for sku, vals in top_gaps
                ],
                "distribution_gap": gap,
            }
        )
    return {"by_state": results}


def calibration_backlog(
    raw_gap: dict[str, dict[str, float]],
    reweighted_gap: dict[str, dict[str, float]],
    *,
    limit: int = 10,
) -> list[dict[str, Any]]:
    """Rank SKU gaps where prospect over/under-index vs buyer (raw and debiased)."""
    items: list[dict[str, Any]] = []
    for token in sorted(set(raw_gap) | set(reweighted_gap)):
        raw = raw_gap.get(token, {})
        rew = reweighted_gap.get(token, {})
        items.append(
            {
                "sku": token,
                "raw_buyer_pct": raw.get("buyer_pct", 0),
                "raw_prospect_pct": raw.get("prospect_pct", 0),
                "raw_gap_points": raw.get("gap_points", 0),
                "reweighted_buyer_pct": rew.get("buyer_pct", 0),
                "reweighted_gap_points": rew.get("gap_points", 0),
                "bias_effect_pp": round(
                    (rew.get("gap_points", 0) or 0) - (raw.get("gap_points", 0) or 0),
                    2,
                ),
            }
        )

    def priority(item: dict[str, Any]) -> float:
        # Prospect over-recommend (negative gap for buyer) = model pushes SKUs buyers don't buy
        return abs(item["reweighted_gap_points"])

    items.sort(key=priority, reverse=True)
    backlog = []
    for item in items[:limit]:
        sku = item["sku"]
        rew_gap = item["reweighted_gap_points"]
        if rew_gap < -5:
            action = "Reduce prospect recommendation share (over-index vs debiased buyer profile)"
        elif rew_gap > 5:
            action = "Increase value/mid-tier recommendation (buyer profile exceeds prospect mix)"
        else:
            action = "Monitor — gap within tolerance"
        backlog.append({**item, "suggested_action": action})
    return backlog
