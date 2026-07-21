#!/usr/bin/env python3
"""Read-only GAP sample: 500 stratified buyer chair purchases vs ORION prospect intel.

Usage (from backend/):
  PYTHONPATH=. python scripts/buyer_gap_sample_500.py [--sample-size 500] [--seed 42]
"""

from __future__ import annotations

import argparse
import csv
import json
import random
import sys
from collections import Counter, defaultdict
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
from dotenv import load_dotenv
from sqlalchemy import create_engine, text

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app.intelligence.buyer_gap_mapping import (
    buyer_compare_sku,
    index_level,
    master_series,
    parse_purchase_token,
    purchase_series,
)
from app.intelligence.product_ladders import resolve_active_ladder

DEFAULT_LEGACY = Path(
    "/Users/josephpark/Desktop/01. Ceragem Consulting/02. Ceragem Dashboard Project/"
    "Ceragem Email Campagin Project/00. Ceragem Purchased Customer Archives/"
    "Ceragem Purchaser Archives 2011-2024.xlsx"
)
DEFAULT_SHOPIFY = Path(
    "/Users/josephpark/Desktop/01. Ceragem Consulting/02. Ceragem Dashboard Project/"
    "Ceragem Email Campagin Project/00. Ceragem Purchased Customer Archives/"
    "Ceragem Purchaser Archives 2024-2026.05.csv"
)


@dataclass
class BuyerRow:
    email: str
    product_raw: str
    sku_token: str
    source: str
    row_id: str


def norm_email(value: str | None) -> str | None:
    email = (value or "").strip().lower()
    if not email or "@" not in email:
        return None
    return email


def load_buyer_rows(legacy: Path, shopify: Path) -> list[BuyerRow]:
    rows: list[BuyerRow] = []
    wb = openpyxl.load_workbook(legacy, read_only=True, data_only=True)
    ws = wb["SALES (202112-202404)"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(headers)}
    for n, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        status = str(row[idx.get("Status", 0)] or "")
        if status.upper() not in ("PAID", "PROCESSING"):
            continue
        email = norm_email(row[idx.get("E MAIL", 0)])
        prod = str(row[idx.get("Material Name", 0)] or "")
        token = parse_purchase_token(prod)
        if email and token:
            rows.append(
                BuyerRow(
                    email=email,
                    product_raw=prod,
                    sku_token=token,
                    source="legacy",
                    row_id=f"legacy:{n}",
                )
            )
    wb.close()

    with open(shopify, newline="", encoding="utf-8-sig") as handle:
        for n, row in enumerate(csv.DictReader(handle), start=2):
            if (row.get("Financial Status") or "").lower() != "paid":
                continue
            email = norm_email(row.get("Email"))
            prod = row.get("Lineitem name") or row.get("Lineitem sku") or ""
            token = parse_purchase_token(prod)
            if email and token:
                rows.append(
                    BuyerRow(
                        email=email,
                        product_raw=prod,
                        sku_token=token,
                        source="shopify",
                        row_id=f"shopify:{n}",
                    )
                )
    return rows


def stratified_sample(
    rows: list[BuyerRow],
    *,
    sample_size: int,
    seed: int,
    priority_row_ids: set[str],
) -> list[BuyerRow]:
    """Include all priority rows (ORION-matched), then fill with stratified random sample."""
    priority = [r for r in rows if r.row_id in priority_row_ids]
    pool = [r for r in rows if r.row_id not in priority_row_ids]
    remaining = max(0, sample_size - len(priority))
    if remaining == 0:
        return priority[:sample_size]

    by_sku: dict[str, list[BuyerRow]] = defaultdict(list)
    for row in pool:
        by_sku[row.sku_token].append(row)

    rng = random.Random(seed)
    total_pool = len(pool)
    picked: list[BuyerRow] = []
    sku_targets = {
        sku: max(1, round(remaining * len(items) / total_pool)) if items else 0
        for sku, items in by_sku.items()
    }
    # Adjust rounding to exact remaining count.
    delta = remaining - sum(sku_targets.values())
    skus_by_weight = sorted(by_sku.keys(), key=lambda s: len(by_sku[s]), reverse=True)
    i = 0
    while delta != 0 and skus_by_weight:
        sku = skus_by_weight[i % len(skus_by_weight)]
        if delta > 0:
            sku_targets[sku] += 1
            delta -= 1
        elif sku_targets.get(sku, 0) > 0:
            sku_targets[sku] -= 1
            delta += 1
        i += 1

    for sku, target in sku_targets.items():
        items = by_sku[sku]
        if not items or target <= 0:
            continue
        picked.extend(rng.sample(items, min(target, len(items))))

    if len(picked) < remaining:
        leftover = [r for r in pool if r not in picked]
        need = remaining - len(picked)
        picked.extend(rng.sample(leftover, min(need, len(leftover))))

    return priority + picked[:remaining]


def load_orion_intel(engine, emails: list[str]) -> dict[str, dict]:
    if not emails:
        return {}
    with engine.connect() as conn:
        rows = conn.execute(
            text(
                """
                SELECT DISTINCT ON (LOWER(TRIM(c.email)))
                       LOWER(TRIM(c.email)) AS email,
                       ci.ceragem_segment,
                       ci.prizm_proxy_segment,
                       ci.purchase_power_index,
                       ci.lifestyle_index,
                       ci.pain_index,
                       ci.recommended_product,
                       c.state
                FROM customers c
                JOIN customer_intelligence ci ON ci.customer_id = c.customer_id
                WHERE LOWER(TRIM(c.email)) = ANY(:emails)
                ORDER BY LOWER(TRIM(c.email)), ci.generated_at DESC NULLS LAST
                """
            ),
            {"emails": emails},
        ).mappings().all()
    return {r["email"]: dict(r) for r in rows}


def prospect_distribution(engine) -> dict[str, float]:
    with engine.connect() as conn:
        rows = conn.execute(
            text(
                """
                SELECT recommended_product, COUNT(*) AS n
                FROM customer_intelligence
                WHERE recommended_product IS NOT NULL
                GROUP BY recommended_product
                """
            )
        ).mappings().all()
    total = sum(r["n"] for r in rows) or 1
    return {r["recommended_product"]: r["n"] / total for r in rows}


def ladder_rank(ladder: list[str], sku: str | None) -> int | None:
    if not sku or sku not in ladder:
        return None
    return ladder.index(sku)


def analyze_row(row: BuyerRow, intel: dict | None) -> dict:
    out = {
        "row_id": row.row_id,
        "email": row.email,
        "source": row.source,
        "purchased_token": row.sku_token,
        "purchased_series": purchase_series(row.sku_token),
        "product_raw": row.product_raw,
        "has_intel": intel is not None,
    }
    if not intel:
        out.update(
            {
                "compare_sku": None,
                "mapping_rule": None,
                "recommended_product": None,
                "exact_hit": None,
                "ladder_hit": None,
                "ladder_rank_gap": None,
                "series_match": None,
                "v_m_cross_gap": None,
            }
        )
        return out

    compare_sku, mapping_rule = buyer_compare_sku(
        row.product_raw,
        ceragem_segment=intel.get("ceragem_segment"),
        prizm_proxy_segment=intel.get("prizm_proxy_segment"),
        purchase_power_index=intel.get("purchase_power_index"),
        lifestyle_index=intel.get("lifestyle_index"),
        pain_index=intel.get("pain_index"),
        customer_state=intel.get("state"),
    )
    recommended = intel.get("recommended_product")
    pain_cat = index_level(intel.get("pain_index"))
    ladder, ladder_source = resolve_active_ladder(
        ceragem_segment=intel.get("ceragem_segment"),
        prizm_segment=intel.get("prizm_proxy_segment"),
        pain_index_category=pain_cat,
    )
    compare_rank = ladder_rank(ladder, compare_sku)
    rec_rank = ladder_rank(ladder, recommended)
    ladder_gap = None
    if compare_rank is not None and rec_rank is not None:
        ladder_gap = compare_rank - rec_rank

    purchased_master = compare_sku
    out.update(
        {
            "compare_sku": compare_sku,
            "mapping_rule": mapping_rule,
            "recommended_product": recommended,
            "ceragem_segment": intel.get("ceragem_segment"),
            "prizm_proxy_segment": intel.get("prizm_proxy_segment"),
            "purchase_power_index": intel.get("purchase_power_index"),
            "lifestyle_index": intel.get("lifestyle_index"),
            "pain_index": intel.get("pain_index"),
            "state": intel.get("state"),
            "ladder_source": ladder_source,
            "exact_hit": compare_sku == recommended if compare_sku and recommended else False,
            "ladder_hit": compare_sku in ladder if compare_sku else False,
            "ladder_rank_gap": ladder_gap,
            "series_match": master_series(purchased_master) == master_series(recommended),
            "v_m_cross_gap": master_series(purchased_master) != master_series(recommended),
            "rec_series": master_series(recommended),
            "compare_series": master_series(purchased_master),
        }
    )
    return out


def pct(n: int, d: int) -> float:
    return round(100.0 * n / d, 2) if d else 0.0


def summarize(records: list[dict], prospect_dist: dict[str, float]) -> dict:
    n = len(records)
    intel_rows = [r for r in records if r["has_intel"]]
    buyer_token_dist = Counter(r["purchased_token"] for r in records)
    buyer_compare_dist = Counter(r["compare_sku"] for r in intel_rows if r.get("compare_sku"))

    by_sku: dict[str, dict] = {}
    for token in sorted(buyer_token_dist.keys()):
        token_rows = [r for r in records if r["purchased_token"] == token]
        token_intel = [r for r in token_rows if r["has_intel"]]
        by_sku[token] = {
            "buyer_rows": len(token_rows),
            "buyer_pct": pct(len(token_rows), n),
            "intel_rows": len(token_intel),
            "exact_hits": sum(1 for r in token_intel if r.get("exact_hit")),
            "exact_hit_rate": pct(sum(1 for r in token_intel if r.get("exact_hit")), len(token_intel)),
            "series_match_rate": pct(sum(1 for r in token_intel if r.get("series_match")), len(token_intel)),
            "v_m_cross_rate": pct(sum(1 for r in token_intel if r.get("v_m_cross_gap")), len(token_intel)),
            "avg_ladder_rank_gap": round(
                sum(r["ladder_rank_gap"] for r in token_intel if r.get("ladder_rank_gap") is not None)
                / max(1, sum(1 for r in token_intel if r.get("ladder_rank_gap") is not None)),
                2,
            ),
            "top_recommended_when_matched": Counter(
                r.get("recommended_product") for r in token_intel if r.get("recommended_product")
            ).most_common(5),
        }

    # Distribution gap: buyer compare SKUs (intel subset) vs prospect recommended mix.
    prospect_token_dist = Counter()
    master_to_token = {
        "Master S4": "V4",
        "Master V5": "V5",
        "Master V6": "V6",
        "Master V7": "V7",
        "Master V9": "V9",
        "Pause M4": "M4",
        "Pause M6": "M6",
        "Pause M6s": "M6S",
        "Pause M10": "M10",
    }
    for master, share in prospect_dist.items():
        token = master_to_token.get(master)
        if token:
            prospect_token_dist[token] += share

    buyer_intel_token = Counter(r["purchased_token"] for r in intel_rows)
    intel_n = len(intel_rows) or 1
    distribution_gap = {}
    all_tokens = sorted(set(buyer_token_dist) | set(prospect_token_dist))
    for token in all_tokens:
        buyer_share = buyer_token_dist.get(token, 0) / n
        prospect_share = prospect_token_dist.get(token, 0.0)
        intel_share = buyer_intel_token.get(token, 0) / intel_n
        distribution_gap[token] = {
            "buyer_sample_pct": round(100 * buyer_share, 2),
            "prospect_recommended_pct": round(100 * prospect_share, 2),
            "gap_points": round(100 * (buyer_share - prospect_share), 2),
            "intel_matched_buyer_pct": round(100 * intel_share, 2),
        }

    ladder_gaps = [r["ladder_rank_gap"] for r in intel_rows if r.get("ladder_rank_gap") is not None]
    return {
        "sample_size": n,
        "intel_matched_rows": len(intel_rows),
        "intel_match_rate_pct": pct(len(intel_rows), n),
        "exact_hit_rows": sum(1 for r in intel_rows if r.get("exact_hit")),
        "exact_hit_rate_pct": pct(sum(1 for r in intel_rows if r.get("exact_hit")), len(intel_rows)),
        "ladder_hit_rows": sum(1 for r in intel_rows if r.get("ladder_hit")),
        "series_match_rows": sum(1 for r in intel_rows if r.get("series_match")),
        "v_m_cross_rows": sum(1 for r in intel_rows if r.get("v_m_cross_gap")),
        "avg_ladder_rank_gap": round(sum(ladder_gaps) / len(ladder_gaps), 2) if ladder_gaps else None,
        "buyer_token_distribution": dict(buyer_token_dist.most_common()),
        "buyer_compare_distribution_intel": dict(buyer_compare_dist.most_common()),
        "distribution_gap_by_sku": distribution_gap,
        "by_purchased_sku": by_sku,
        "mapping_rules": Counter(r.get("mapping_rule") for r in intel_rows if r.get("mapping_rule")).most_common(),
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="500-sample buyer GAP analysis")
    parser.add_argument("--sample-size", type=int, default=500)
    parser.add_argument("--seed", type=int, default=42)
    parser.add_argument("--legacy", type=Path, default=DEFAULT_LEGACY)
    parser.add_argument("--shopify", type=Path, default=DEFAULT_SHOPIFY)
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path(__file__).resolve().parents[1] / "reports",
    )
    args = parser.parse_args()

    load_dotenv(Path(__file__).resolve().parents[1] / ".env")
    import os

    engine = create_engine(os.environ["DATABASE_URL"])

    all_rows = load_buyer_rows(args.legacy, args.shopify)
    all_emails = sorted({r.email for r in all_rows})
    intel_map = load_orion_intel(engine, all_emails)
    matched_row_ids = {r.row_id for r in all_rows if r.email in intel_map}

    sample = stratified_sample(
        all_rows,
        sample_size=args.sample_size,
        seed=args.seed,
        priority_row_ids=matched_row_ids,
    )
    prospect_dist = prospect_distribution(engine)
    records = [analyze_row(r, intel_map.get(r.email)) for r in sample]
    summary = summarize(records, prospect_dist)

    args.output_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    json_path = args.output_dir / f"buyer_gap_sample_{args.sample_size}_{stamp}.json"
    csv_path = args.output_dir / f"buyer_gap_sample_{args.sample_size}_{stamp}.csv"

    payload = {
        "generated_at": stamp,
        "sample_size": args.sample_size,
        "seed": args.seed,
        "total_buyer_chair_rows": len(all_rows),
        "total_orion_matched_emails": len(intel_map),
        "priority_matched_rows_in_sample": sum(1 for r in sample if r.row_id in matched_row_ids),
        "summary": summary,
        "records": records,
    }
    json_path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")

    fieldnames = [
        "row_id",
        "email",
        "source",
        "purchased_token",
        "purchased_series",
        "has_intel",
        "compare_sku",
        "mapping_rule",
        "recommended_product",
        "exact_hit",
        "ladder_hit",
        "ladder_rank_gap",
        "series_match",
        "v_m_cross_gap",
        "ceragem_segment",
        "prizm_proxy_segment",
        "purchase_power_index",
        "lifestyle_index",
        "state",
    ]
    with open(csv_path, "w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(records)

    print(json.dumps({"json": str(json_path), "csv": str(csv_path), "summary": summary}, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
